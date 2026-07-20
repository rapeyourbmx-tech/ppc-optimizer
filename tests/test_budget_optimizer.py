"""Tests for campaign budget optimization."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.config import ThresholdConfiguration, load_configuration
from app.models.budget import BudgetAction
from app.services.budget_optimizer import BudgetOptimizer
from app.services.multi_campaign_analyzer import MultiCampaignAnalyzer
from main import run


def _write_campaign(path: Path, rows: list[dict]) -> None:
    """Write one campaign report file from row dictionaries."""
    pd.DataFrame(rows).to_csv(path, index=False)


def _product(sku: str, cost: float, conversions: float, value: float) -> dict:
    """Build one product row."""
    return {
        "Item ID": sku,
        "Item title": f"Product {sku}",
        "Impressions": 1000,
        "Clicks": 20,
        "Cost": cost,
        "Conversions": conversions,
        "Conversion Value": value,
    }


@pytest.fixture
def optimizer_configuration() -> ThresholdConfiguration:
    """Configuration with explicit budget thresholds for the tests."""
    return ThresholdConfiguration.model_validate(
        {
            "budget": {
                "increase_efficiency": 5,
                "decrease_efficiency": 1,
                "shift_share": 0.2,
                "confidence_conversions": 10,
            }
        }
    )


@pytest.fixture
def two_campaign_report(tmp_path: Path, optimizer_configuration: ThresholdConfiguration):
    """A strong campaign and a weak campaign analyzed together."""
    strong_path = tmp_path / "strong.csv"
    weak_path = tmp_path / "weak.csv"
    # Strong: SCALE product -> high growth share and ROAS.
    _write_campaign(
        strong_path,
        [_product("STRONG-1", cost=400.0, conversions=8.0, value=8000.0)],
    )
    # Weak: converting but saturated (no SCALE, above watch threshold).
    _write_campaign(
        weak_path,
        [_product("WEAK-1", cost=1000.0, conversions=2.0, value=500.0)],
    )
    return MultiCampaignAnalyzer(configuration=optimizer_configuration).analyze(
        [strong_path, weak_path]
    )


def test_optimize_assesses_every_campaign(
    two_campaign_report,
    optimizer_configuration: ThresholdConfiguration,
) -> None:
    """Every campaign gets spend, ROAS, efficiency, and saturation scores."""
    report = BudgetOptimizer(optimizer_configuration).optimize(two_campaign_report)

    strong, weak = report.assessments
    assert strong.current_spend == 400.0
    assert strong.current_budget == strong.current_spend
    assert strong.current_roas == 20.0
    assert strong.marginal_efficiency == 20.0
    assert strong.saturation == 0.0
    assert weak.current_roas == 0.5
    assert weak.saturation == 1.0
    assert 0.0 <= weak.marginal_efficiency <= 0.5


def test_optimize_selects_actions_from_config_thresholds(
    two_campaign_report,
    optimizer_configuration: ThresholdConfiguration,
) -> None:
    """Actions follow the configured efficiency thresholds."""
    report = BudgetOptimizer(optimizer_configuration).optimize(two_campaign_report)

    strong, weak = report.assessments
    assert strong.action is BudgetAction.INCREASE
    assert weak.action is BudgetAction.DECREASE


def test_optimize_builds_transfer_with_gain_and_confidence(
    two_campaign_report,
    optimizer_configuration: ThresholdConfiguration,
) -> None:
    """A transfer moves the configured share to the best campaign."""
    report = BudgetOptimizer(optimizer_configuration).optimize(two_campaign_report)

    assert len(report.transfers) == 1
    transfer = report.transfers[0]
    assert transfer.source_campaign == "weak"
    assert transfer.destination_campaign == "strong"
    assert transfer.amount == pytest.approx(200.0)
    assert transfer.expected_revenue_increase == pytest.approx(200.0 * 20.0)
    assert transfer.confidence == pytest.approx(0.8)
    assert report.total_expected_gain == transfer.expected_revenue_increase


def test_optimize_without_decrease_campaigns_moves_nothing(
    tmp_path: Path,
    optimizer_configuration: ThresholdConfiguration,
) -> None:
    """Balanced portfolios produce no transfers."""
    first_path = tmp_path / "first.csv"
    second_path = tmp_path / "second.csv"
    _write_campaign(
        first_path,
        [_product("A-1", cost=400.0, conversions=8.0, value=8000.0)],
    )
    _write_campaign(
        second_path,
        [_product("B-1", cost=400.0, conversions=8.0, value=8000.0)],
    )
    combined = MultiCampaignAnalyzer(configuration=optimizer_configuration).analyze(
        [first_path, second_path]
    )

    report = BudgetOptimizer(optimizer_configuration).optimize(combined)

    assert report.transfers == []
    assert report.total_expected_gain == 0.0


def test_run_writes_budget_section_and_action_plan(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The workbook contains the budget table and the action plan."""
    strong_path = tmp_path / "strong.csv"
    weak_path = tmp_path / "weak.csv"
    _write_campaign(
        strong_path,
        [_product("STRONG-1", cost=400.0, conversions=8.0, value=8000.0)],
    )
    _write_campaign(
        weak_path,
        [_product("WEAK-1", cost=1000.0, conversions=2.0, value=500.0)],
    )
    config_path = tmp_path / "config.yaml"
    config_path.write_text(
        "budget:\n"
        "  increase_efficiency: 5\n"
        "  decrease_efficiency: 1\n"
        "  shift_share: 0.2\n"
        "  confidence_conversions: 10\n",
        encoding="utf-8",
    )
    output_path = tmp_path / "report.xlsx"

    exit_code = run(
        [strong_path, weak_path],
        config_path=config_path,
        output_path=output_path,
    )

    assert exit_code == 0
    workbook = load_workbook(output_path)
    dashboard_values = [
        cell.value
        for row in workbook["Dashboard"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    summary_values = [
        cell.value
        for row in workbook["Executive Summary"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]

    assert "Budget Optimization" in dashboard_values
    assert "DECREASE" in dashboard_values
    assert "INCREASE" in dashboard_values
    assert "Expected total gain" in dashboard_values
    assert "Action Plan" in summary_values
    assert any(value.startswith("• Decrease weak by 200.00") for value in summary_values)
    assert any(value.startswith("• Increase strong by 200.00") for value in summary_values)
    assert any("Expected monthly gain: +4,000.00 revenue" in value for value in summary_values)
