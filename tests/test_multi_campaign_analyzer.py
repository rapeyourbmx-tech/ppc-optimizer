"""Tests for multi-campaign report analysis."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.services.multi_campaign_analyzer import MultiCampaignAnalyzer
from main import run


def _write_report(path: Path, sku: str, cost: float, conversion_value: float) -> None:
    """Write one minimal single-product report file."""
    pd.DataFrame(
        [
            {
                "Item ID": sku,
                "Item title": f"Product {sku}",
                "Impressions": 1000,
                "Clicks": 30,
                "Cost": cost,
                "Conversions": 1.0 if conversion_value else 0.0,
                "Conversion Value": conversion_value,
            }
        ]
    ).to_csv(path, index=False)


@pytest.fixture
def campaign_files(tmp_path: Path) -> list[Path]:
    """Create three campaign report files with distinct outcomes."""
    high_path = tmp_path / "high.csv"
    average_path = tmp_path / "average.csv"
    low_path = tmp_path / "low.csv"
    _write_report(high_path, "HIGH-1", cost=350.0, conversion_value=49095.0)
    _write_report(average_path, "AVG-1", cost=450.0, conversion_value=0.0)
    _write_report(low_path, "LOW-1", cost=45.0, conversion_value=0.0)
    return [high_path, average_path, low_path]


def test_analyze_builds_one_campaign_report_per_file(
    campaign_files: list[Path],
) -> None:
    """Every source file becomes a campaign with derived metadata."""
    report = MultiCampaignAnalyzer().analyze(campaign_files)

    assert [campaign.metadata.name for campaign in report.campaigns] == [
        "high",
        "average",
        "low",
    ]
    assert [campaign.metadata.campaign_type for campaign in report.campaigns] == [
        "High priority",
        "Medium priority",
        "Low priority",
    ]
    assert report.campaigns[0].metadata.source_file == "high.csv"
    assert all(
        campaign.campaign_summary.total_products == 1
        for campaign in report.campaigns
    )


def test_analyze_combines_products_with_campaign_columns(
    campaign_files: list[Path],
) -> None:
    """The combined frame keeps campaign identity for every product row."""
    report = MultiCampaignAnalyzer().analyze(campaign_files)

    assert list(report.products.columns[:3]) == [
        "campaign_name",
        "campaign_type",
        "source_file",
    ]
    assert len(report.products) == 3
    assert len(report.decisions) == 3
    assert report.products.loc[0, "campaign_name"] == "high"
    assert report.products.loc[1, "source_file"] == "average.csv"


def test_analyze_aggregates_overall_summary(campaign_files: list[Path]) -> None:
    """The overall summary totals every campaign's metrics."""
    report = MultiCampaignAnalyzer().analyze(campaign_files)
    overall = report.overall_summary

    assert overall.total_products == 3
    assert overall.total_cost == 845.0
    assert overall.total_revenue == 49095.0
    assert overall.total_conversions == 1.0
    assert (overall.keep, overall.watch, overall.pause, overall.scale) == (0, 1, 1, 1)
    assert report.overall_health == "Needs attention"


def test_analyze_rejects_empty_source_list() -> None:
    """Analyzing zero files is an explicit error."""
    with pytest.raises(ValueError, match="At least one report file"):
        MultiCampaignAnalyzer().analyze([])


def test_run_with_multiple_files_prints_per_campaign_and_overall_summary(
    campaign_files: list[Path],
    tmp_path: Path,
    capsys: pytest.CaptureFixture[str],
) -> None:
    """The CLI reports every campaign and the combined totals."""
    output_path = tmp_path / "combined.xlsx"

    exit_code = run(campaign_files, output_path=output_path)

    captured = capsys.readouterr()
    assert exit_code == 0
    assert "high: Health:" in captured.out
    assert "Overall: Health: Needs attention | Products: 3" in captured.out
    assert output_path.is_file()


def test_workbook_contains_campaign_comparison_and_recommendations(
    campaign_files: list[Path],
    tmp_path: Path,
) -> None:
    """The dashboard compares campaigns; the summary lists per-campaign advice."""
    output_path = tmp_path / "combined.xlsx"
    run(campaign_files, output_path=output_path)

    workbook = load_workbook(output_path)
    dashboard_values = [
        cell.value
        for row in workbook["Dashboard"].iter_rows()
        for cell in row
        if cell.value is not None
    ]
    summary_values = [
        cell.value
        for row in workbook["Executive Summary"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]

    assert "Campaign Comparison" in dashboard_values
    assert "high" in dashboard_values
    assert any(
        isinstance(value, str) and value.startswith("=COUNTIF(") and "high" in value
        for value in dashboard_values
    )
    assert "Recommended actions per campaign" in summary_values
    assert any("average (Medium priority, average.csv)" in value for value in summary_values)
