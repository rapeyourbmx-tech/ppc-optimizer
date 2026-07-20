"""Tests for the Excel workbook exporter."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.core.workbook import WorkbookSheet
from app.reporting.excel_workbook_exporter import ExcelWorkbookExporter
from app.services.multi_campaign_analyzer import MultiCampaignAnalyzer


@pytest.fixture
def exported_workbook_path(tmp_path: Path) -> Path:
    """Run the pipeline on a small report and export the workbook."""
    source_path = tmp_path / "product_report.csv"
    pd.DataFrame(
        [
            {
                "Item ID": "SCALE-1",
                "Item title": "Scaling product",
                "Impressions": 5000,
                "Clicks": 30,
                "Cost": 350.0,
                "Conversions": 1.0,
                "Conversion Value": 49095.0,
            },
            {
                "Item ID": "PAUSE-1",
                "Item title": "Wasting product",
                "Impressions": 1000,
                "Clicks": 104,
                "Cost": 450.0,
                "Conversions": 0.0,
                "Conversion Value": 0.0,
            },
            {
                "Item ID": "WATCH-1",
                "Item title": "Young product",
                "Impressions": 200,
                "Clicks": 3,
                "Cost": 45.0,
                "Conversions": 0.0,
                "Conversion Value": 0.0,
            },
        ]
    ).to_csv(source_path, index=False)

    report = MultiCampaignAnalyzer().analyze([source_path])
    output_path = tmp_path / "report.xlsx"
    ExcelWorkbookExporter().export(report, output_path)
    return output_path


def test_export_creates_every_required_sheet(exported_workbook_path: Path) -> None:
    """The workbook contains exactly the required sheets in order."""
    workbook = load_workbook(exported_workbook_path)

    assert workbook.sheetnames == [str(sheet) for sheet in WorkbookSheet]


def test_products_sheet_has_decision_columns_and_layout(
    exported_workbook_path: Path,
) -> None:
    """The Products sheet appends decision columns and sets layout features."""
    sheet = load_workbook(exported_workbook_path)[str(WorkbookSheet.PRODUCTS)]
    headers = [cell.value for cell in sheet[1]]

    assert headers[-4:] == ["Status", "ROAS", "Recommendation", "Reason"]
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref is not None
    status_column = headers.index("Status") + 1
    statuses = {
        sheet.cell(row=row_number, column=status_column).value
        for row_number in range(2, 5)
    }
    assert statuses == {"SCALE", "PAUSE", "WATCH"}


def test_products_sheet_roas_is_a_formula(exported_workbook_path: Path) -> None:
    """ROAS cells recalculate from cost and conversion value."""
    sheet = load_workbook(exported_workbook_path)[str(WorkbookSheet.PRODUCTS)]
    headers = [cell.value for cell in sheet[1]]
    roas_column = headers.index("ROAS") + 1

    roas_value = sheet.cell(row=2, column=roas_column).value

    assert isinstance(roas_value, str)
    assert roas_value.startswith("=IFERROR(")


def test_status_sheets_contain_matching_products(exported_workbook_path: Path) -> None:
    """Each status sheet lists only the products with that status."""
    workbook = load_workbook(exported_workbook_path)

    assert workbook[str(WorkbookSheet.SCALE)].cell(row=2, column=1).value == "SCALE-1"
    assert workbook[str(WorkbookSheet.PAUSE)].cell(row=2, column=1).value == "PAUSE-1"
    assert workbook[str(WorkbookSheet.WATCH)].cell(row=2, column=1).value == "WATCH-1"
    keep_note = workbook[str(WorkbookSheet.KEEP)].cell(row=2, column=1).value
    assert keep_note == "No products with the KEEP status."


def test_top_sheets_rank_winners_and_losers(exported_workbook_path: Path) -> None:
    """Top Winners lists revenue products; Top Losers lists wasted spend."""
    workbook = load_workbook(exported_workbook_path)

    assert workbook[str(WorkbookSheet.TOP_WINNERS)].cell(row=2, column=1).value == "SCALE-1"
    assert workbook[str(WorkbookSheet.TOP_LOSERS)].cell(row=2, column=1).value == "PAUSE-1"


def test_dashboard_kpis_are_formula_backed(exported_workbook_path: Path) -> None:
    """Dashboard KPI values are formulas over the Products sheet."""
    sheet = load_workbook(exported_workbook_path)[str(WorkbookSheet.DASHBOARD)]
    formulas = [
        cell.value
        for row in sheet.iter_rows()
        for cell in row
        if isinstance(cell.value, str) and cell.value.startswith("=")
    ]

    assert any("COUNTA(Products!" in formula for formula in formulas)
    assert any('COUNTIF(' in formula and '"KEEP"' in formula for formula in formulas)
    assert any("SUMIFS(" in formula and '"PAUSE"' in formula for formula in formulas)


def test_exporter_applies_excel_and_dashboard_settings(tmp_path: Path) -> None:
    """Custom font, dashboard title, and top-list size come from configuration."""
    from app.config import ThresholdConfiguration

    source_path = tmp_path / "product_report.csv"
    pd.DataFrame(
        [
            {
                "Item ID": f"WIN-{index}",
                "Item title": f"Winner {index}",
                "Impressions": 5000,
                "Clicks": 30,
                "Cost": 350.0,
                "Conversions": 3.0,
                "Conversion Value": 40000.0 + index,
            }
            for index in range(3)
        ]
    ).to_csv(source_path, index=False)
    configuration = ThresholdConfiguration.model_validate(
        {
            "excel": {"font_name": "Calibri", "top_list_size": 1},
            "dashboard": {"title": "Custom Title"},
        }
    )
    report = MultiCampaignAnalyzer(configuration=configuration).analyze([source_path])
    output_path = tmp_path / "report.xlsx"

    ExcelWorkbookExporter(configuration).export(report, output_path)

    workbook = load_workbook(output_path)
    dashboard = workbook[str(WorkbookSheet.DASHBOARD)]
    winners = workbook[str(WorkbookSheet.TOP_WINNERS)]
    assert dashboard.cell(row=1, column=2).value == "Custom Title"
    assert dashboard.cell(row=1, column=2).font.name == "Calibri"
    assert winners.cell(row=2, column=1).value is not None
    assert winners.cell(row=3, column=1).value is None
