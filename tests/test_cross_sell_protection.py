"""Tests for the cross-sell revenue protection logic and reporting."""

from pathlib import Path

import pandas as pd
import pytest
from openpyxl import load_workbook

from app.analyzers.decision_explainer import CROSS_SELL_PROTECTION_REASON
from app.analyzers.product_analyzer import ProductAnalyzer
from app.config import CrossSellThresholds, DecisionThresholds
from app.models.product_decision import ProductStatus
from main import run


def _analyze(**overrides):
    """Analyze one product with pause-triggering defaults."""
    metrics = {
        "sku": "PROTECT-1",
        "clicks": 30,
        "cost": 500.0,
        "conversions": 0.0,
        "conversion_value": 0.0,
        "assist_revenue": 0.0,
    }
    metrics.update(overrides)
    return ProductAnalyzer().analyze(**metrics)


def test_pause_candidate_with_strong_cross_sell_roas_becomes_watch() -> None:
    """A cross-sell ROAS above the threshold protects the pause candidate."""
    decision = _analyze(assist_revenue=2464.48)  # 2464.48 / 500 = 4.93x

    assert decision.status is ProductStatus.WATCH
    assert decision.reason == CROSS_SELL_PROTECTION_REASON
    assert decision.assist_revenue == 2464.48


def test_cross_sell_roas_below_threshold_keeps_pause() -> None:
    """Weak cross-sell revenue no longer blocks the PAUSE recommendation."""
    decision = _analyze(assist_revenue=500.0)  # 500 / 500 = 1.0x < 2.0

    assert decision.status is ProductStatus.PAUSE


def test_cross_sell_roas_equal_to_threshold_becomes_watch() -> None:
    """The protection is inclusive at exactly the configured ratio."""
    decision = _analyze(assist_revenue=1000.0)  # 1000 / 500 = 2.0x

    assert decision.status is ProductStatus.WATCH


def test_disabled_protection_keeps_pause() -> None:
    """cross_sell.enabled: false turns the protection off entirely."""
    analyzer = ProductAnalyzer(
        thresholds=DecisionThresholds(cross_sell=CrossSellThresholds(enabled=False))
    )

    decision = analyzer.analyze(
        sku="OFF-1",
        clicks=30,
        cost=500.0,
        conversions=0.0,
        conversion_value=0.0,
        assist_revenue=9999.0,
    )

    assert decision.status is ProductStatus.PAUSE


def test_pause_candidate_without_assist_revenue_still_pauses() -> None:
    """The PAUSE rule itself is unchanged when there is no assist revenue."""
    decision = _analyze()

    assert decision.status is ProductStatus.PAUSE


def test_assist_revenue_does_not_change_scale_decisions() -> None:
    """SCALE stays SCALE and its ROAS ignores assist revenue."""
    with_assist = _analyze(conversions=3.0, conversion_value=8000.0, assist_revenue=9999.0)
    without_assist = _analyze(conversions=3.0, conversion_value=8000.0)

    assert with_assist.status is ProductStatus.SCALE
    assert with_assist.status == without_assist.status
    assert with_assist.roas == without_assist.roas


def test_assist_revenue_does_not_change_keep_decisions() -> None:
    """KEEP stays KEEP regardless of assist revenue."""
    with_assist = _analyze(conversions=2.0, conversion_value=900.0, assist_revenue=5000.0)

    assert with_assist.status is ProductStatus.KEEP


def test_roas_uses_conversion_value_only() -> None:
    """Assist revenue never enters the ROAS calculation."""
    decision = _analyze(conversions=2.0, conversion_value=1000.0, assist_revenue=9000.0)

    assert decision.roas == pytest.approx(200.0)


def test_protection_explanation_shows_ratio_threshold_and_reason() -> None:
    """Explain mode shows the amount, the ratio, the threshold, and the reason."""
    decision = _analyze(assist_revenue=2464.48)

    assert "Cross-sell revenue = 2464.48" in decision.explanation
    assert "Cross-sell ROAS = 4.93" in decision.explanation
    assert "Threshold = 2.00" in decision.explanation
    assert CROSS_SELL_PROTECTION_REASON in decision.explanation


def test_workbook_reports_protection_and_top_assist_sheet(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    """The workbook carries the protection block, assist sheet, and winner filter."""
    monkeypatch.chdir(tmp_path)
    source_path = tmp_path / "statistics.csv"
    pd.DataFrame(
        [
            {
                "Offer ID": "PROTECTED",
                "Impressions": 5000,
                "Clicks": 40,
                "Cost": 500.0,
                "Conversions": 0.0,
                "Direct Revenue": 0.0,
                "Cross-sell Revenue": 2464.48,
            },
            {
                "Offer ID": "WINNER",
                "Impressions": 5000,
                "Clicks": 40,
                "Cost": 400.0,
                "Conversions": 2.0,
                "Direct Revenue": 8000.0,
                "Cross-sell Revenue": 100.0,
            },
            {
                "Offer ID": "NO-SALES",
                "Impressions": 5000,
                "Clicks": 40,
                "Cost": 350.0,
                "Conversions": 0.0,
                "Direct Revenue": 500.0,
                "Cross-sell Revenue": 0.0,
            },
        ]
    ).to_csv(source_path, index=False)
    output_path = tmp_path / "report.xlsx"

    exit_code = run(source_path, output_path=output_path)

    assert exit_code == 0
    workbook = load_workbook(output_path)

    summary_values = [
        cell.value
        for row in workbook["Executive Summary"].iter_rows()
        for cell in row
        if isinstance(cell.value, str)
    ]
    assert "Cross-sell Protection" in summary_values
    assert "Products protected by Cross-sell Revenue: 1" in summary_values
    assert any(
        value.startswith("Total Cross-sell Revenue protected: 2,464.48") for value in summary_values
    )

    assist_sheet = workbook["Top Assist Products"]
    assert [cell.value for cell in assist_sheet[1]][:2] == ["SKU", "Product"]
    assert assist_sheet.cell(row=2, column=1).value == "PROTECTED"
    assert assist_sheet.cell(row=2, column=7).value == 2464.48
    assert assist_sheet.cell(row=3, column=1).value == "WINNER"

    winner_skus = [
        row[0].value for row in workbook["Top Winners"].iter_rows(min_row=2) if row[0].value
    ]
    assert "WINNER" in winner_skus
    assert "NO-SALES" not in winner_skus
