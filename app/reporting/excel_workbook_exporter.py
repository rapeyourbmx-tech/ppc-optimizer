"""Excel workbook export for analyzed Google Ads product reports."""

from dataclasses import dataclass
from datetime import date
from pathlib import Path

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule, FormulaRule
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.workbook import WorkbookSheet
from app.models.budget import BudgetAction, BudgetOptimizationReport
from app.models.campaign import CampaignReport, MultiCampaignReport
from app.models.product_decision import ProductDecision, ProductStatus

_FONT_NAME = "Arial"
_HEADER_FILL = PatternFill("solid", fgColor="1F3864")
_HEADER_FONT = Font(name=_FONT_NAME, bold=True, color="FFFFFF", size=11)
_TITLE_FONT = Font(name=_FONT_NAME, bold=True, size=16, color="1F3864")
_SUBTITLE_FONT = Font(name=_FONT_NAME, size=10, color="808080")
_BODY_FONT = Font(name=_FONT_NAME, size=11)
_NOTE_FONT = Font(name=_FONT_NAME, size=11, italic=True, color="808080")
_THIN_BORDER = Border(*[Side(style="thin", color="D9D9D9")] * 4)

_FORMAT_MONEY = "#,##0.00"
_FORMAT_INTEGER = "#,##0"
_FORMAT_PERCENT = "#,##0%"
_FORMAT_CONVERSIONS = "0.00"
_FORMAT_CTR = '0.00"%"'

_BUDGET_ACTION_COLORS: dict[str, str] = {
    BudgetAction.INCREASE: "006100",
    BudgetAction.KEEP: "808080",
    BudgetAction.DECREASE: "9C0006",
}
_STATUS_STYLES: dict[str, tuple[str, str]] = {
    ProductStatus.KEEP: ("C6EFCE", "006100"),
    ProductStatus.WATCH: ("FFEB9C", "9C6500"),
    ProductStatus.PAUSE: ("FFC7CE", "9C0006"),
    ProductStatus.SCALE: ("BDD7EE", "1F4E79"),
}
_RECOMMENDATIONS: dict[str, str] = {
    ProductStatus.KEEP: "Keep running at the current budget.",
    ProductStatus.WATCH: "Monitor performance and collect more data.",
    ProductStatus.PAUSE: "Pause the product to stop wasted spend.",
    ProductStatus.SCALE: "Increase the budget and bids for this product.",
}
_PRETTY_COLUMN_NAMES: dict[str, str] = {
    "campaign_name": "Campaign",
    "campaign_type": "Campaign Type",
    "source_file": "Source File",
    "product_id": "Product ID",
    "product": "Product",
    "clicks": "Clicks",
    "impressions": "Impressions",
    "ctr": "CTR",
    "average_cpc": "Avg. CPC",
    "cost": "Cost",
    "conversions": "Conversions",
    "conversion_value": "Conversion Value",
    "cost_per_conversion": "Cost / Conv.",
    "all_conversions": "All Conversions",
    "all_conversion_value": "All Conv. Value",
}
_COLUMN_FORMATS: dict[str, str] = {
    "clicks": _FORMAT_INTEGER,
    "impressions": _FORMAT_INTEGER,
    "ctr": _FORMAT_CTR,
    "average_cpc": _FORMAT_MONEY,
    "cost": _FORMAT_MONEY,
    "conversions": _FORMAT_CONVERSIONS,
    "conversion_value": _FORMAT_MONEY,
    "cost_per_conversion": _FORMAT_MONEY,
    "all_conversions": _FORMAT_CONVERSIONS,
    "all_conversion_value": _FORMAT_MONEY,
}
_TOP_LIST_SIZE = 10
_MAX_COLUMN_WIDTH = 45.0
_MIN_COLUMN_WIDTH = 9.0


@dataclass(frozen=True, slots=True)
class _ProductsLayout:
    """Cell coordinates of key Products-sheet columns for formulas."""

    sku_column: str
    campaign_column: str
    cost_column: str
    revenue_column: str
    conversions_column: str
    status_column: str
    first_row: int
    last_row: int

    def column_range(self, column_letter: str) -> str:
        """Return an absolute Products-sheet range for one column."""
        return (
            f"Products!${column_letter}${self.first_row}:"
            f"${column_letter}${self.last_row}"
        )


class ExcelWorkbookExporter:
    """Export an analyzed product report to a formatted Excel workbook."""

    def export(
        self,
        report: MultiCampaignReport,
        output_path: Path,
        budget: BudgetOptimizationReport | None = None,
    ) -> None:
        """Write the complete report workbook to the supplied path."""
        workbook = Workbook()
        workbook.remove(workbook.active)

        products_sheet = workbook.create_sheet(WorkbookSheet.PRODUCTS)
        layout = self._write_products_sheet(products_sheet, report)

        dashboard_sheet = workbook.create_sheet(WorkbookSheet.DASHBOARD, index=0)
        self._write_dashboard_sheet(dashboard_sheet, report, layout, budget)

        summary_sheet = workbook.create_sheet(WorkbookSheet.EXECUTIVE_SUMMARY, index=1)
        self._write_executive_summary_sheet(summary_sheet, report, layout, budget)

        for status in (
            ProductStatus.KEEP,
            ProductStatus.WATCH,
            ProductStatus.PAUSE,
            ProductStatus.SCALE,
        ):
            status_sheet = workbook.create_sheet(str(status))
            self._write_status_sheet(status_sheet, report, status)

        winners_sheet = workbook.create_sheet(WorkbookSheet.TOP_WINNERS)
        self._write_top_products_sheet(winners_sheet, report, winners=True)
        losers_sheet = workbook.create_sheet(WorkbookSheet.TOP_LOSERS)
        self._write_top_products_sheet(losers_sheet, report, winners=False)

        workbook.save(output_path)

    # ------------------------------------------------------------------
    # Products sheet
    # ------------------------------------------------------------------
    def _write_products_sheet(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
    ) -> _ProductsLayout:
        """Write every original column plus decision columns, return the layout."""
        sheet.sheet_properties.tabColor = "1F3864"
        products = report.products
        source_columns = list(products.columns)
        headers = [
            *(self._pretty_column_name(column) for column in source_columns),
            "Status",
            "ROAS",
            "Recommendation",
            "Reason",
        ]
        self._write_header_row(sheet, headers)

        cost_index = source_columns.index("cost") + 1
        revenue_index = source_columns.index("conversion_value") + 1
        conversions_index = source_columns.index("conversions") + 1
        status_index = len(source_columns) + 1
        roas_index = status_index + 1
        recommendation_index = roas_index + 1
        reason_index = recommendation_index + 1
        cost_letter = get_column_letter(cost_index)
        revenue_letter = get_column_letter(revenue_index)

        for row_offset, (row_values, decision) in enumerate(
            zip(products.itertuples(index=False), report.decisions, strict=True)
        ):
            row_number = row_offset + 2
            for column_offset, (column_name, value) in enumerate(
                zip(source_columns, row_values, strict=True)
            ):
                cell = sheet.cell(row=row_number, column=column_offset + 1)
                cell.value = None if pd.isna(value) else value
                cell.font = _BODY_FONT
                cell.number_format = _COLUMN_FORMATS.get(column_name, "General")

            status_cell = sheet.cell(row=row_number, column=status_index)
            status_cell.value = str(decision.status)
            fill_color, font_color = _STATUS_STYLES[decision.status]
            status_cell.fill = PatternFill("solid", fgColor=fill_color)
            status_cell.font = Font(name=_FONT_NAME, bold=True, color=font_color)
            status_cell.alignment = Alignment(horizontal="center")

            roas_cell = sheet.cell(row=row_number, column=roas_index)
            roas_cell.value = (
                f"=IFERROR({revenue_letter}{row_number}/{cost_letter}{row_number},0)"
            )
            roas_cell.number_format = _FORMAT_PERCENT
            roas_cell.font = _BODY_FONT

            recommendation_cell = sheet.cell(row=row_number, column=recommendation_index)
            recommendation_cell.value = _RECOMMENDATIONS[decision.status]
            recommendation_cell.font = _BODY_FONT
            reason_cell = sheet.cell(row=row_number, column=reason_index)
            reason_cell.value = decision.reason
            reason_cell.font = _BODY_FONT

        last_row = len(report.decisions) + 1
        last_letter = get_column_letter(len(headers))
        status_letter = get_column_letter(status_index)
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:{last_letter}{last_row}"
        self._apply_status_row_formatting(
            sheet,
            data_range=f"A2:{last_letter}{last_row}",
            status_letter=status_letter,
        )
        if last_row > 1:
            roas_letter = get_column_letter(roas_index)
            sheet.conditional_formatting.add(
                f"{roas_letter}2:{roas_letter}{last_row}",
                _roas_color_scale(),
            )
            sheet.conditional_formatting.add(
                f"{cost_letter}2:{cost_letter}{last_row}",
                _cost_data_bar(),
            )
        self._auto_fit_columns(sheet, formula_width=10.0)

        return _ProductsLayout(
            sku_column=get_column_letter(source_columns.index("product_id") + 1),
            campaign_column=get_column_letter(source_columns.index("campaign_name") + 1),
            cost_column=cost_letter,
            revenue_column=revenue_letter,
            conversions_column=get_column_letter(conversions_index),
            status_column=status_letter,
            first_row=2,
            last_row=max(last_row, 2),
        )

    # ------------------------------------------------------------------
    # Dashboard sheet
    # ------------------------------------------------------------------
    def _write_dashboard_sheet(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        layout: _ProductsLayout,
        budget: BudgetOptimizationReport | None,
    ) -> None:
        """Write KPI cards backed by formulas over the Products sheet."""
        sheet.sheet_properties.tabColor = "1F3864"
        sheet.sheet_view.showGridLines = False
        self._write_title(
            sheet,
            title="PPC Optimizer — Campaign Dashboard",
            subtitle=(
                f"Generated {date.today().isoformat()} · "
                f"{report.overall_summary.total_products} products · "
                f"{len(report.campaigns)} campaign(s)"
            ),
        )

        sku_range = layout.column_range(layout.sku_column)
        cost_range = layout.column_range(layout.cost_column)
        revenue_range = layout.column_range(layout.revenue_column)
        conversions_range = layout.column_range(layout.conversions_column)
        status_range = layout.column_range(layout.status_column)

        top_cards = (
            ("TOTAL PRODUCTS", f"=COUNTA({sku_range})", _FORMAT_INTEGER, "1F3864"),
            ("TOTAL COST", f"=SUM({cost_range})", _FORMAT_MONEY, "1F3864"),
            ("TOTAL REVENUE", f"=SUM({revenue_range})", _FORMAT_MONEY, "1F3864"),
            (
                "OVERALL ROAS",
                f"=IFERROR(SUM({revenue_range})/SUM({cost_range}),0)",
                _FORMAT_PERCENT,
                "1F3864",
            ),
        )
        status_cards = tuple(
            (
                str(status),
                f'=COUNTIF({status_range},"{status}")',
                _FORMAT_INTEGER,
                _STATUS_STYLES[status][1],
            )
            for status in (
                ProductStatus.KEEP,
                ProductStatus.WATCH,
                ProductStatus.PAUSE,
                ProductStatus.SCALE,
            )
        )
        bottom_cards = (
            (
                "WASTED SPEND (PAUSE)",
                f'=SUMIFS({cost_range},{status_range},"PAUSE")',
                _FORMAT_MONEY,
                "9C0006",
            ),
            (
                "SCALE REVENUE",
                f'=SUMIFS({revenue_range},{status_range},"SCALE")',
                _FORMAT_MONEY,
                "1F4E79",
            ),
            ("TOTAL CONVERSIONS", f"=SUM({conversions_range})", _FORMAT_CONVERSIONS, "1F3864"),
            ("CAMPAIGN HEALTH", report.overall_health, "General", "9C6500"),
        )

        for row_anchor, cards in ((4, top_cards), (8, status_cards), (12, bottom_cards)):
            for card_offset, (card_title, card_value, number_format, accent) in enumerate(cards):
                self._write_kpi_card(
                    sheet,
                    anchor_row=row_anchor,
                    anchor_column=2 + card_offset * 4,
                    title=card_title,
                    value=card_value,
                    number_format=number_format,
                    accent_color=accent,
                )

        self._write_campaign_comparison(sheet, report, layout, anchor_row=16)
        if budget is not None:
            budget_anchor = 16 + len(report.campaigns) + 3
            self._write_budget_optimization(sheet, report, layout, budget, budget_anchor)

        for column_index in range(1, 19):
            sheet.column_dimensions[get_column_letter(column_index)].width = 12.0
        sheet.column_dimensions["B"].width = 24.0
        sheet.column_dimensions["I"].width = 16.0

    def _write_campaign_comparison(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        layout: _ProductsLayout,
        *,
        anchor_row: int,
    ) -> None:
        """Write the formula-backed campaign comparison table."""
        section_title = sheet.cell(row=anchor_row, column=2, value="Campaign Comparison")
        section_title.font = Font(name=_FONT_NAME, bold=True, size=12, color="1F3864")

        headers = ("Campaign", "Products", "Cost", "Revenue", "ROAS", "Scale", "Pause", "Health")
        header_row = anchor_row + 1
        self._write_header_row(sheet, headers, row=header_row, start_column=2)
        campaign_range = layout.column_range(layout.campaign_column)
        cost_range = layout.column_range(layout.cost_column)
        revenue_range = layout.column_range(layout.revenue_column)
        status_range = layout.column_range(layout.status_column)

        for campaign_offset, campaign in enumerate(report.campaigns):
            row_number = header_row + 1 + campaign_offset
            escaped_name = campaign.metadata.name.replace('"', '""')
            name_criterion = f'"{escaped_name}"'
            values: tuple[object, ...] = (
                campaign.metadata.name,
                f"=COUNTIF({campaign_range},{name_criterion})",
                f"=SUMIFS({cost_range},{campaign_range},{name_criterion})",
                f"=SUMIFS({revenue_range},{campaign_range},{name_criterion})",
                f"=IFERROR(E{row_number}/D{row_number},0)",
                (
                    f'=COUNTIFS({campaign_range},{name_criterion},'
                    f'{status_range},"SCALE")'
                ),
                (
                    f'=COUNTIFS({campaign_range},{name_criterion},'
                    f'{status_range},"PAUSE")'
                ),
                campaign.health,
            )
            formats = (
                "General",
                _FORMAT_INTEGER,
                _FORMAT_MONEY,
                _FORMAT_MONEY,
                _FORMAT_PERCENT,
                _FORMAT_INTEGER,
                _FORMAT_INTEGER,
                "General",
            )
            for column_offset, (value, number_format) in enumerate(
                zip(values, formats, strict=True)
            ):
                cell = sheet.cell(row=row_number, column=2 + column_offset)
                cell.value = value
                cell.number_format = number_format
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER

    def _write_budget_optimization(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        layout: _ProductsLayout,
        budget: BudgetOptimizationReport,
        anchor_row: int,
    ) -> None:
        """Write the budget redistribution section of the dashboard."""
        section_title = sheet.cell(row=anchor_row, column=2, value="Budget Optimization")
        section_title.font = Font(name=_FONT_NAME, bold=True, size=12, color="1F3864")

        headers = (
            "Campaign",
            "Current Spend",
            "ROAS",
            "Recommendation",
            "Move Budget",
            "Expected Gain",
        )
        header_row = anchor_row + 1
        self._write_header_row(sheet, headers, row=header_row, start_column=2)

        campaign_range = layout.column_range(layout.campaign_column)
        cost_range = layout.column_range(layout.cost_column)
        revenue_range = layout.column_range(layout.revenue_column)
        moved_out = {
            transfer.source_campaign: transfer.amount for transfer in budget.transfers
        }
        moved_in: dict[str, float] = {}
        gains: dict[str, float] = {}
        for transfer in budget.transfers:
            moved_in[transfer.destination_campaign] = (
                moved_in.get(transfer.destination_campaign, 0.0) + transfer.amount
            )
            gains[transfer.source_campaign] = transfer.expected_revenue_increase

        for assessment_offset, assessment in enumerate(budget.assessments):
            row_number = header_row + 1 + assessment_offset
            escaped_name = assessment.campaign_name.replace('"', '""')
            name_criterion = f'"{escaped_name}"'
            movement = moved_in.get(assessment.campaign_name, 0.0) - moved_out.get(
                assessment.campaign_name, 0.0
            )
            values: tuple[object, ...] = (
                assessment.campaign_name,
                f"=SUMIFS({cost_range},{campaign_range},{name_criterion})",
                (
                    f"=IFERROR(SUMIFS({revenue_range},{campaign_range},{name_criterion})"
                    f"/SUMIFS({cost_range},{campaign_range},{name_criterion}),0)"
                ),
                str(assessment.action),
                movement,
                gains.get(assessment.campaign_name, 0.0),
            )
            formats = (
                "General",
                _FORMAT_MONEY,
                _FORMAT_PERCENT,
                "General",
                '+#,##0.00;-#,##0.00;"—"',
                '+#,##0.00;-#,##0.00;"—"',
            )
            for column_offset, (value, number_format) in enumerate(
                zip(values, formats, strict=True)
            ):
                cell = sheet.cell(row=row_number, column=2 + column_offset)
                cell.value = value
                cell.number_format = number_format
                cell.font = _BODY_FONT
                cell.border = _THIN_BORDER
            action_cell = sheet.cell(row=row_number, column=5)
            action_cell.font = Font(
                name=_FONT_NAME,
                bold=True,
                color=_BUDGET_ACTION_COLORS[assessment.action],
            )

        total_row = header_row + 1 + len(budget.assessments)
        total_label = sheet.cell(row=total_row, column=2, value="Expected total gain")
        total_label.font = Font(name=_FONT_NAME, bold=True)
        total_cell = sheet.cell(row=total_row, column=7, value=budget.total_expected_gain)
        total_cell.number_format = '+#,##0.00;-#,##0.00;"—"'
        total_cell.font = Font(name=_FONT_NAME, bold=True, color="006100")

    def _write_kpi_card(
        self,
        sheet: Worksheet,
        *,
        anchor_row: int,
        anchor_column: int,
        title: str,
        value: object,
        number_format: str,
        accent_color: str,
    ) -> None:
        """Render one three-column KPI card with a title and a large value."""
        end_column = anchor_column + 2
        title_cell = sheet.cell(row=anchor_row, column=anchor_column)
        title_cell.value = title
        title_cell.font = Font(name=_FONT_NAME, size=9, bold=True, color="808080")
        title_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(
            start_row=anchor_row,
            start_column=anchor_column,
            end_row=anchor_row,
            end_column=end_column,
        )

        value_cell = sheet.cell(row=anchor_row + 1, column=anchor_column)
        value_cell.value = value
        value_cell.number_format = number_format
        value_cell.font = Font(name=_FONT_NAME, size=18, bold=True, color=accent_color)
        value_cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.merge_cells(
            start_row=anchor_row + 1,
            start_column=anchor_column,
            end_row=anchor_row + 2,
            end_column=end_column,
        )

        card_fill = PatternFill("solid", fgColor="F5F7FA")
        for row_number in range(anchor_row, anchor_row + 3):
            for column_number in range(anchor_column, end_column + 1):
                cell = sheet.cell(row=row_number, column=column_number)
                cell.border = _THIN_BORDER
                cell.fill = card_fill

    # ------------------------------------------------------------------
    # Executive Summary sheet
    # ------------------------------------------------------------------
    def _write_executive_summary_sheet(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        layout: _ProductsLayout,
        budget: BudgetOptimizationReport | None,
    ) -> None:
        """Write the narrative summary with formula-backed key metrics."""
        sheet.sheet_properties.tabColor = "808080"
        sheet.sheet_view.showGridLines = False
        self._write_title(
            sheet,
            title="Executive Summary",
            subtitle=f"Generated {date.today().isoformat()}",
        )

        health_label = sheet.cell(row=4, column=2, value="Campaign health")
        health_label.font = Font(name=_FONT_NAME, bold=True, size=12)
        health_value = sheet.cell(row=4, column=4, value=report.overall_health)
        health_value.font = Font(name=_FONT_NAME, bold=True, size=12, color="9C6500")

        cost_range = layout.column_range(layout.cost_column)
        revenue_range = layout.column_range(layout.revenue_column)
        status_range = layout.column_range(layout.status_column)
        metrics: tuple[tuple[str, object, str], ...] = (
            ("Products analyzed", f"=COUNTA({layout.column_range(layout.sku_column)})", _FORMAT_INTEGER),
            ("Total advertising cost", f"=SUM({cost_range})", _FORMAT_MONEY),
            ("Total conversion value", f"=SUM({revenue_range})", _FORMAT_MONEY),
            (
                "Overall ROAS",
                f"=IFERROR(SUM({revenue_range})/SUM({cost_range}),0)",
                _FORMAT_PERCENT,
            ),
            (
                "Wasted spend (PAUSE products)",
                f'=SUMIFS({cost_range},{status_range},"PAUSE")',
                _FORMAT_MONEY,
            ),
            (
                "Revenue from SCALE products",
                f'=SUMIFS({revenue_range},{status_range},"SCALE")',
                _FORMAT_MONEY,
            ),
        )
        metrics_start_row = 6
        for metric_offset, (label, value, number_format) in enumerate(metrics):
            row_number = metrics_start_row + metric_offset
            label_cell = sheet.cell(row=row_number, column=2, value=label)
            label_cell.font = _BODY_FONT
            value_cell = sheet.cell(row=row_number, column=4, value=value)
            value_cell.font = Font(name=_FONT_NAME, bold=True)
            value_cell.number_format = number_format
            value_cell.alignment = Alignment(horizontal="right")

        campaigns_needing_attention = sum(
            1 for campaign in report.campaigns if campaign.health != "Healthy"
        )
        summary_row = metrics_start_row + len(metrics) + 1
        summary_cell = sheet.cell(row=summary_row, column=2)
        summary_cell.value = (
            f"{campaigns_needing_attention} of {len(report.campaigns)} "
            "campaign(s) need attention."
        )
        summary_cell.font = _NOTE_FONT

        actions_row = summary_row + 2
        actions_title = sheet.cell(
            row=actions_row,
            column=2,
            value="Recommended actions per campaign",
        )
        actions_title.font = Font(name=_FONT_NAME, bold=True, size=12)
        current_row = actions_row + 1
        for campaign in report.campaigns:
            current_row = self._write_campaign_recommendations(sheet, campaign, current_row)
        if budget is not None:
            self._write_action_plan(sheet, report, budget, current_row + 1)

        sheet.column_dimensions["A"].width = 3.0
        for column_letter in ("B", "C", "D", "E", "F", "G", "H"):
            sheet.column_dimensions[column_letter].width = 18.0

    def _write_action_plan(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        budget: BudgetOptimizationReport,
        anchor_row: int,
    ) -> None:
        """Write the budget redistribution action plan."""
        title_cell = sheet.cell(row=anchor_row, column=2, value="Action Plan")
        title_cell.font = Font(name=_FONT_NAME, bold=True, size=12)
        currency = self._detect_currency(report)
        currency_suffix = f" {currency}" if currency else ""

        current_row = anchor_row + 1
        if not budget.transfers:
            note_cell = sheet.cell(row=current_row, column=2)
            note_cell.value = (
                "• No budget redistribution recommended — "
                "the current allocation is balanced."
            )
            note_cell.font = _BODY_FONT
            return

        for transfer in budget.transfers:
            decrease_cell = sheet.cell(row=current_row, column=2)
            decrease_cell.value = (
                f"• Decrease {transfer.source_campaign} by "
                f"{transfer.amount:,.2f}{currency_suffix}"
            )
            decrease_cell.font = _BODY_FONT
            increase_cell = sheet.cell(row=current_row + 1, column=2)
            increase_cell.value = (
                f"• Increase {transfer.destination_campaign} by "
                f"{transfer.amount:,.2f}{currency_suffix} "
                f"(confidence {transfer.confidence:.0%})"
            )
            increase_cell.font = _BODY_FONT
            current_row += 2

        gain_cell = sheet.cell(row=current_row + 1, column=2)
        gain_cell.value = (
            f"Expected monthly gain: +{budget.total_expected_gain:,.2f} revenue"
        )
        gain_cell.font = Font(name=_FONT_NAME, bold=True, color="006100")

    @staticmethod
    def _detect_currency(report: MultiCampaignReport) -> str:
        """Return the dominant currency code from the source data, if any."""
        for column_name in ("код_валюти", "currency_code", "currency"):
            if column_name in report.products.columns:
                values = report.products[column_name].dropna().astype(str)
                if not values.empty:
                    return str(values.mode().iloc[0])

        return ""

    @staticmethod
    def _write_campaign_recommendations(
        sheet: Worksheet,
        campaign: CampaignReport,
        start_row: int,
    ) -> int:
        """Write one campaign's recommendation block, return the next free row."""
        metadata = campaign.metadata
        title_cell = sheet.cell(row=start_row + 1, column=2)
        title_cell.value = (
            f"{metadata.name} ({metadata.campaign_type}, {metadata.source_file}) — "
            f"{campaign.health}"
        )
        title_cell.font = Font(name=_FONT_NAME, bold=True, color="1F3864")

        recommendations = campaign.report.audit_report.recommendations
        for recommendation_offset, recommendation in enumerate(recommendations, start=2):
            cell = sheet.cell(row=start_row + recommendation_offset, column=2)
            cell.value = f"• {recommendation}"
            cell.font = _BODY_FONT

        return start_row + len(recommendations) + 2

    # ------------------------------------------------------------------
    # Per-status and top-products sheets
    # ------------------------------------------------------------------
    def _write_status_sheet(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        status: ProductStatus,
    ) -> None:
        """Write every decision of one status into its own sheet."""
        fill_color, font_color = _STATUS_STYLES[status]
        sheet.sheet_properties.tabColor = fill_color
        decisions = [
            row
            for row in self._decision_rows(report)
            if row[0].status is status
        ]
        self._write_decision_table(
            sheet,
            rows=decisions,
            empty_message=f"No products with the {status} status.",
        )

    def _write_top_products_sheet(
        self,
        sheet: Worksheet,
        report: MultiCampaignReport,
        *,
        winners: bool,
    ) -> None:
        """Write the top winners (by ROAS) or top losers (by wasted cost)."""
        paired = self._decision_rows(report)
        if winners:
            sheet.sheet_properties.tabColor = "006100"
            selected = sorted(
                (pair for pair in paired if pair[0].conversion_value > 0),
                key=lambda pair: pair[0].roas,
                reverse=True,
            )
            empty_message = "No products with conversion value yet."
        else:
            sheet.sheet_properties.tabColor = "9C0006"
            selected = sorted(
                (
                    pair
                    for pair in paired
                    if pair[0].conversions == 0 and pair[0].cost > 0
                ),
                key=lambda pair: pair[0].cost,
                reverse=True,
            )
            empty_message = "No products with spend and zero conversions."
        self._write_decision_table(
            sheet,
            rows=selected[:_TOP_LIST_SIZE],
            empty_message=empty_message,
        )

    def _write_decision_table(
        self,
        sheet: Worksheet,
        *,
        rows: list[tuple[ProductDecision, str, str]],
        empty_message: str,
    ) -> None:
        """Write one decision table with shared columns and formatting."""
        headers = (
            "SKU",
            "Campaign",
            "Product",
            "Clicks",
            "Cost",
            "Conversions",
            "Revenue",
            "ROAS",
            "Reason",
        )
        self._write_header_row(sheet, headers)
        column_formats = (
            "General",
            "General",
            "General",
            _FORMAT_INTEGER,
            _FORMAT_MONEY,
            _FORMAT_CONVERSIONS,
            _FORMAT_MONEY,
            _FORMAT_PERCENT,
            "General",
        )

        for row_offset, (decision, product_name, campaign_name) in enumerate(rows):
            row_number = row_offset + 2
            values: tuple[object, ...] = (
                decision.sku,
                campaign_name,
                product_name,
                decision.clicks,
                decision.cost,
                decision.conversions,
                decision.conversion_value,
                f"=IFERROR(G{row_number}/E{row_number},0)",
                decision.reason,
            )
            for column_offset, (value, number_format) in enumerate(
                zip(values, column_formats, strict=True)
            ):
                cell = sheet.cell(row=row_number, column=column_offset + 1)
                cell.value = value
                cell.number_format = number_format
                cell.font = _BODY_FONT

        last_row = len(rows) + 1
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = f"A1:I{max(last_row, 1)}"
        if rows:
            sheet.conditional_formatting.add(f"H2:H{last_row}", _roas_color_scale())
            sheet.conditional_formatting.add(f"E2:E{last_row}", _cost_data_bar())
        else:
            note_cell = sheet.cell(row=2, column=1, value=empty_message)
            note_cell.font = _NOTE_FONT
        self._auto_fit_columns(sheet, formula_width=10.0)

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------
    @staticmethod
    def _decision_rows(
        report: MultiCampaignReport,
    ) -> list[tuple[ProductDecision, str, str]]:
        """Pair every decision with its product and campaign names."""
        products = report.products
        if "product" in products.columns:
            product_names = products["product"].fillna("").astype(str).tolist()
        else:
            product_names = [""] * len(products)
        campaign_names = products["campaign_name"].astype(str).tolist()

        return list(zip(report.decisions, product_names, campaign_names, strict=True))

    @staticmethod
    def _pretty_column_name(column_name: str) -> str:
        """Return the display header for one normalized source column."""
        if column_name in _PRETTY_COLUMN_NAMES:
            return _PRETTY_COLUMN_NAMES[column_name]

        return column_name.replace("_", " ").strip().capitalize()

    @staticmethod
    def _write_header_row(
        sheet: Worksheet,
        headers: tuple[str, ...] | list[str],
        *,
        row: int = 1,
        start_column: int = 1,
    ) -> None:
        """Write one styled header row."""
        for column_offset, header in enumerate(headers):
            cell = sheet.cell(row=row, column=start_column + column_offset, value=header)
            cell.fill = _HEADER_FILL
            cell.font = _HEADER_FONT
            cell.alignment = Alignment(horizontal="center", vertical="center")
        sheet.row_dimensions[row].height = 22.0

    @staticmethod
    def _write_title(sheet: Worksheet, *, title: str, subtitle: str) -> None:
        """Write the sheet title and subtitle rows."""
        title_cell = sheet.cell(row=1, column=2, value=title)
        title_cell.font = _TITLE_FONT
        subtitle_cell = sheet.cell(row=2, column=2, value=subtitle)
        subtitle_cell.font = _SUBTITLE_FONT

    @staticmethod
    def _apply_status_row_formatting(
        sheet: Worksheet,
        *,
        data_range: str,
        status_letter: str,
    ) -> None:
        """Tint whole data rows according to the Status column value."""
        for status, (fill_color, _font_color) in _STATUS_STYLES.items():
            sheet.conditional_formatting.add(
                data_range,
                FormulaRule(
                    formula=[f'${status_letter}2="{status}"'],
                    fill=PatternFill("solid", fgColor=_lighten(fill_color)),
                ),
            )

    @staticmethod
    def _auto_fit_columns(sheet: Worksheet, *, formula_width: float) -> None:
        """Size every column to its longest value within sensible bounds."""
        for column_cells in sheet.columns:
            lengths = [formula_width]
            column_letter = None
            for cell in column_cells:
                column_letter = column_letter or cell.column_letter
                if cell.value is None:
                    continue
                text = str(cell.value)
                if text.startswith("="):
                    continue
                lengths.append(float(len(text)) + 2.0)
            if column_letter is not None:
                width = min(max(max(lengths), _MIN_COLUMN_WIDTH), _MAX_COLUMN_WIDTH)
                sheet.column_dimensions[column_letter].width = width


def _roas_color_scale() -> ColorScaleRule:
    """Return the red-to-green color scale used for ROAS columns."""
    return ColorScaleRule(
        start_type="num",
        start_value=0,
        start_color="F8696B",
        mid_type="num",
        mid_value=1,
        mid_color="FFEB84",
        end_type="num",
        end_value=10,
        end_color="63BE7B",
    )


def _cost_data_bar() -> DataBarRule:
    """Return the data bar used for cost columns."""
    return DataBarRule(
        start_type="min",
        end_type="max",
        color="638EC6",
        showValue=True,
    )


def _lighten(hex_color: str) -> str:
    """Return a lighter tint of a six-digit hex color for row fills."""
    red, green, blue = (int(hex_color[index : index + 2], 16) for index in (0, 2, 4))
    lightened = tuple(int(channel + (255 - channel) * 0.6) for channel in (red, green, blue))
    return "".join(f"{channel:02X}" for channel in lightened)
